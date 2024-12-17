from opensearchpy import OpenSearch, helpers
import psycopg2
from datetime import datetime
import uuid
import ast
from concurrent.futures import ThreadPoolExecutor, as_completed
import os
import pandas as pd
import boto3
import psycopg2
import io
import concurrent.futures
import logging
import time
from io import BytesIO
from datetime import datetime
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
import threading  # For asynchronous execution
import time
from io import StringIO
import json
from decimal import Decimal
from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import Alignment, PatternFill
from openpyxl.utils import get_column_letter
import re
import pytz
from pytz import timezone

from common_utils.db_utils import DB
db_config = {
    'host': "amoppostoct19.c3qae66ke1lg.us-east-1.rds.amazonaws.com",
    'port': "5432",
    'user':"root",
    'password':"AmopTeam123"
}

# Connect to OpenSearch
es = OpenSearch(
    ['https://search-amopsearch-df66xuwugs7f6b43ihav5yd5zm.us-east-1.es.amazonaws.com'],
    http_auth=('admin', 'Amopteam@123'),
    use_ssl=True,
    verify_certs=True,
    timeout=60
)

# Connect to PostgreSQL
# conn = psycopg2.connect(
#     dbname="altaworx_central",
#     user="root",
#     password="AmopTeam123",
#     host="amoppostoct19.c3qae66ke1lg.us-east-1.rds.amazonaws.com",
#     port="5432"
# )

def get_table_schema(conn, table_name):
    query = """
        SELECT column_name, data_type
        FROM information_schema.columns
        WHERE table_name = %s
        ORDER BY ordinal_position
    """
    with conn.cursor() as cur:
        cur.execute(query, (table_name,))
        columns = cur.fetchall()

    return {column[0]: column[1] for column in columns}

def get_ranges(elements, range_size=20):
    elements = sorted(elements)  # Ensure the list is sorted
    ranges = []

    for i in range(0, len(elements), range_size):
        start = elements[i]
        end = elements[min(i + range_size - 1, len(elements) - 1)]
        ranges.append(f"{start}-{end}" if start != end else str(start))

    return ranges

def parse_datetime(value):
    if value in (None, 'null', 'NULL', ''):
        return None
    if isinstance(value, str):
        try:
            return datetime.fromisoformat(value).isoformat()
        except ValueError:
            try:
                return datetime.strptime(value, '%Y-%m-%d %H:%M:%S.%f').isoformat()
            except ValueError:
                print(f"Warning: Unable to parse date time string: {value}\n")
                return value
    elif isinstance(value, datetime):
        return value.isoformat()
    return value

def convert_value(value, data_type):
    if value in (None, 'null', 'NULL', ''):
        return None
    if data_type == 'boolean':
        return str(value).lower() in ('true', '1')
    elif data_type == 'timestamp without time zone':
        return parse_datetime(value)
    elif data_type == ('character varying', 'text'):
        return str(value)
    elif data_type == 'integer':
        return int(value)
    elif data_type == 'real':
        return float(value)
    elif data_type == 'uuid':
        try:
            return str(uuid.UUID(value))
        except ValueError:
            print(f"Warning: Invalid UUID format: {value}\n")
            return value
    elif data_type == 'json':
        if isinstance(value, list):
            return value  # Directly return if the value is already a list
        try:
            return json.loads(value)
        except json.JSONDecodeError:
            print(f"Warning: Invalid JSON format: {value}\n")
            return value
    return value

def bulk_index_data(batch_data):
    table_name, rows, schema = batch_data
    actions = []
    for item in rows:
        doc = {}
        for i, column_name in enumerate(schema.keys()):
            value = item[i]
            data_type = schema[column_name]
            converted_value = convert_value(value, data_type)
            doc[column_name] = converted_value

        action = {
            "_op_type": "index",  # Use "update" if you want to update existing documents
            "_index": table_name,
            "_source": doc,
            "_id": doc.get("id")
        }
        actions.append(action)

    try:
        response = helpers.bulk(es, actions)
        print(f"Successfully indexed {response[0]} documents for {table_name}")
    except Exception as e:
        print(f"Error indexing documents: {e}")

def fetch_and_bulk_index_data(table_name):
    schema = get_table_schema(conn, table_name)
    columns_list = ', '.join([f'"{col}"' for col in schema.keys()])
    query = f"SELECT {columns_list} FROM {table_name} "

    with conn.cursor() as cur:
        cur.execute(query)
        # print(rows)
        while True:
            rows = cur.fetchmany(BATCH_SIZE)
            batch_count=batch_count+1
            if not rows:
                break
            yield (table_name, rows, schema)

def fetch_table_names():
    table_query = """
        select search_tables from open_search_index;
        """
    try:
        with conn.cursor() as cur:
            cur.execute(table_query)
            tables= cur.fetchall()
            # Extract table names from the query results
            print("tables",tables)
    except psycopg2.Error as e:
        print(f"Error fetching table names: {e}")

    index_tables = []
    for item in tables:
        # Check if the item is a tuple with a single string element
        if isinstance(item, tuple) and len(item) == 1:
            value = item[0]
            if value.startswith("{") and value.endswith("}"):
                # Remove curly braces and split by comma
                value = value.strip("{}").replace("'", "")
                tables = [table.strip() for table in value.split(",")]
                index_tables.extend(tables)
            else:
                index_tables.append(value)

    index_tables = list(set(index_tables))
    #index_tables=["optimization_group"]
    print("---------------",index_tables)
    return index_tables

def fetch_data_with_pagination(table_name, batch_size, start_row=0):
    schema = get_table_schema(conn, table_name)
    columns_list = ', '.join([f'"{col}"' for col in schema.keys()])
    query = f"SELECT {columns_list} FROM {table_name} LIMIT %s OFFSET %s"

    with conn.cursor() as cur:
        while True:
            cur.execute(query, (batch_size, start_row))
            rows = cur.fetchall()
            if not rows:
                break
            yield (table_name, rows, schema)
            start_row += batch_size

def process_table(table_name):
    batch_size = 25000
    for batch_data in fetch_data_with_pagination(table_name, batch_size):
        print(F"got data")
        # table_name, rows, schema = batch_data
    # batches=fetch_and_bulk_index_data(table_name)
    # for i,batch_data in enumerate(batches):
    #     print(f"got batch {i}")
        bulk_index_data(batch_data)


def get_list_view_cols(conn):
    query = """
        SELECT db_column_name FROM field_column_mapping where module_name = 'Customer Rate Plan' and table_col = 'yes'
    """
    with conn.cursor() as cur:
        cur.execute(query)
        # columns = cur.fetchall()
        # print("-------------------",columns)

        columns = [row[0] for row in cur.fetchall()]
        return columns
def filter_boolean_fields(fields, field_types):
    """
    Filter out boolean fields from the list of fields to be searched.
    """
    return [field for field in fields if field_types.get(field) != 'boolean']

def remove_field(columns, field_to_remove):
    """
    Remove a specific field from the list of columns if it exists.
    """
    return [col for col in columns if col != field_to_remove]

def get_field_types(index_name):
    try:
        # Retrieve the mapping for the index
        mapping = es.indices.get_mapping(index=index_name)
        # Extract the mapping of the properties/fields
        properties = mapping[index_name]['mappings']['properties']

        field_types = {}
        for field, field_data in properties.items():
            field_types[field] = field_data.get('type', 'unknown')

        return field_types

    except Exception as e:
        print(f"An error occurred while retrieving field types: {e}")
        return None
def filter_supported_fields(columns, field_types):
    """
    Filters columns to include only those that support prefix queries (keyword or text).
    """
    return [col for col in columns if field_types.get(col) in ['keyword', 'text']]


def reindex_all():
    table_names = fetch_table_names()
    NUM_THREADS=4

    # Use multiprocessing Pool to parallelize the indexing process
    # with Pool(processes=NUM_PROCESSES) as pool:
    #     pool.map(process_table, table_names)
    with ThreadPoolExecutor(max_workers=NUM_THREADS) as executor:
        futures = [executor.submit(process_table, table) for table in table_names]
    print("Tejaswi**************")
    #cursor.close()
    conn.close()

    return {'index_Status': 'Indexing for all Docs '}

# def elasticsearch_indexing():
#     index_tables=fetch_table_names()
#     print("---------------",index_tables)
#     for table in index_tables:
#         fetch_and_bulk_index_data(table)


#def search_data(query, search_all_columns,index_name,):
# def search_data(query, search_all_columns, index_name, start, end,db_name):
#     if db_name == "altaworx_central":
#         index=index_name
#     else:
#         index=f"{index_name}_{db_name}"
#     print("index_name is ----------------",index)





#     search_body = {
#         "query": {
#             "bool": {
#                 "must": [
#                     {
#                         "query_string": {
#                             "query": f"{query}*",
#                             "fields": ["*"]  # Search across all fields
#                         }
#                     }
#                 ]
#             }
#         },
#         "from": start,  # Calculate the offset
#         "size": end-start  # Number of results per page
#         }



#     if index_name == "sim_management_inventory_action_history" or index_name == "sim_management_bulk_change_request":
#         custom_filter={}
#     elif index_name == "sim_management_inventory":
#         custom_filter={"is_active": True, "tenant_id": 1 }
#     else:
#         custom_filter={"is_active": True}
#     filter_clause=create_filter_clause(index_name,custom_filter)
#     if filter_clause:
#         search_body["query"]["bool"]["filter"] = filter_clause
#     try:
#         response = es.search(index=index, body=search_body)
#         cleaned_response = replace_none_with_empty(response)
#         return cleaned_response
#     except Exception as e:
#         print(f"An error occurred while searching: {e}")
#         return None

def create_filter_clause(index_name, custom_filter):
    filter_clauses = []

    # Check if a custom_filter dictionary is provided
    if custom_filter:
        for field, value in custom_filter.items():
            # Only include fields that exist in the index
            if is_field_present(index_name, field):
                filter_clauses.append({"term": {field: value}})

    # If no filters were added, return None
    if not filter_clauses:
        return None

    # Return the filter clauses
    return filter_clauses



def search_data(query, search_all_columns, index_name,start,end,columns,db_name, custom_filter):

    if db_name == "altaworx_central":
        index=index_name
    else:
        index=f"{index_name}_{db_name}"
    print("+++++++++++++++",columns)
    columns = [col for col in columns if col and col != "is_active"]
    columns=list(set(columns))

    print("************",columns)


    search_body = {
        "query": {
            "bool": {
                "must": [
                    {
                         "multi_match": {
                            "query": f"{query}",  # Flexible query input
                            "fields": columns,  # Search across all specified fields
                            "lenient": True,  # Treats fields as a single field
                            "type": "phrase"

                        }
                    }
                ]
            }
        },
        "from": start,  # Calculate the offset
        "size": end-start  # Number of results per page
    }

    # if index_name == "sim_management_inventory_action_history" or index_name == "sim_management_bulk_change_request":
    #     custom_filter={}
    # elif index_name == "sim_management_inventory":
    #     custom_filter={"is_active": True, "tenant_id": 1 }
    # else:
    #     custom_filter={"is_active": True}
    filter_clause=create_filter_clause(index_name,custom_filter)
    if filter_clause:
        search_body["query"]["bool"]["filter"] = filter_clause
    try:
        response = es.search(index=index, body=search_body)
        cleaned_response = replace_none_with_empty(response)
        return cleaned_response
    except Exception as e:
        print(f"An error occurred while searching: {e}")
        return None




def create_filter_clause(index_name, custom_filter):
    filter_clauses = []

    # Check if a custom_filter dictionary is provided
    if custom_filter:
        for field, value in custom_filter.items():
            # Only include fields that exist in the index
            if is_field_present(index_name, field):
                filter_clauses.append({"term": {field: value}})

    # If no filters were added, return None
    if not filter_clauses:
        return None

    # Return the filter clauses
    return filter_clauses

def get_field_type(index_name, field_name):
    """Fetch the type of a field from the OpenSearch index mapping."""
    try:
        # Fetch the index mapping
        print("--------------",index_name , field_name)
        mapping = es.indices.get_mapping(index=index_name)

        # Navigate to the properties of the index
        properties = mapping[index_name]['mappings']['properties']

        # Check if the field exists
        if field_name in properties:
            return properties[field_name].get('type', 'unknown')  # Return the field type
        else:
            return None  # Field does not exist
    except Exception as e:
        print(f"Error fetching field type: {e}")
        return None

def is_field_present(index_name, field_name):
    try:
        # Get the mapping for the index
        mapping = es.indices.get_mapping(index=index_name)
        # Check if the field exists in the mapping
        fields = mapping[index_name]['mappings']['properties']
        return field_name in fields
    except Exception as e:
        print(f"An error occurred while checking field presence: {e}")
        return False

def convert_filter_values(filters, index_name):
    print("-----------",index_name)
    converted_filters = {}
    for field_name, values in filters.items():
        field_type = get_field_type(index_name, field_name)

        # Convert values based on field type
        if field_type == "integer":
            converted_filters[field_name] = [int(v) for v in values]
        elif field_type == "float":
            converted_filters[field_name] = [float(v) for v in values]
        elif field_type == "keyword":
            converted_filters[field_name] = [str(v) for v in values]
        else:
            converted_filters[field_name] = values
    return converted_filters

def get_field_type(index_name, field_name):
    """
    Fetch the field type from the OpenSearch index mappings.
    """
    try:
        mapping = es.indices.get_mapping(index=index_name)
        field_mapping = mapping.get(index_name, {}).get("mappings", {}).get("properties", {})
        return field_mapping.get(field_name, {}).get("type", None)
    except Exception as e:
        print(f"Error fetching mapping for index {index_name}: {e}")
        return None

def convert_date_format(dates, column_name="billing_cycle_end_date", current_format="%m/%d/%Y %H:%M:%S", target_format="%Y-%m-%dT%H:%M:%S"):
    if isinstance(dates, list):
        formatted_dates = []
        for date_str in dates:
            try:
                date_obj = datetime.strptime(date_str, current_format)
                if column_name == "billing_cycle_end_date":
                    if date_obj.time() == pd.Timestamp("23:59:59").time():
                        date_obj = (date_obj + pd.Timedelta(days=1)).replace(hour=0, minute=0, second=0, microsecond=0)
                formatted_date = date_obj.strftime(target_format)
                formatted_dates.append(formatted_date)
            except ValueError as e:
                raise ValueError(f"Error parsing date '{date_str}': {e}")
        return formatted_dates
    else:
        try:
            date_obj = datetime.strptime(dates, current_format)
            if column_name == "billing_cycle_end_date":
                if date_obj.time() == pd.Timestamp("23:59:59").time():
                    date_obj = (date_obj + pd.Timedelta(days=1)).replace(hour=0, minute=0, second=0, microsecond=0)
            return date_obj.strftime(target_format)
        except ValueError as e:
            raise ValueError(f"Error parsing date '{dates}': {e}")

def advance_search_data(filters,index_list,start, end,db_name):
    if db_name=="altaworx_central":
            index_list=index_list
    else:
        index_list=f"{index_list}_{db_name}"
    print("index_name------",index_list)

    search_body = {
        "query": {
            "bool": {
                "must": []
            }
        },
        "from": start ,  # Start from the specified offset
        "size": end-start,
         "track_total_hits": True # Number of results to return
    }

    # Add filters to the search body
    # if filters:
    #     for field, values in filters.items():
    #         if values:  # Check if values is not empty
    #             if len(values) == 1:
    #                 # Single value, use term query
    #                 search_body["query"]["bool"]["must"].append({
    #                     "term": {
    #                         f"{field}.keyword": values[0]
    #                     }
    #                 })
    #             else:
    #                 # Multiple values, use terms query
    #                 search_body["query"]["bool"]["must"].append({
    #                     "terms": {
    #                          f"{field}.keyword": values
    #                     }
    #                 })
    # if filters:
    #     for field, values in filters.items():
    #         if values:  # Check if values is not empty
    #             if isinstance(values, list):  # Handle lists
    #                 if len(values) == 1:
    #                     # Single value, handle boolean or other types
    #                     value = values[0]
    #                     print("***********************8",values, "-----------",type(values))
    #                     if isinstance(value, bool):  # Handle boolean condition
    #                         search_body["query"]["bool"]["must"].append({
    #                             "term": {f"{field}": value}
    #                         })
    #                     elif isinstance(values, float):  # Single integer condition
    #                         search_body["query"]["bool"]["must"].append({
    #                             "term": {f"{field}": values}
    #                         })
    #                     elif isinstance(values, str):  # Single integer condition
    #                         search_body["query"]["bool"]["must"].append({
    #                             "term": {f"{field}": values}
    #                         })
    #                     elif field == 'bulk_change_id':  # Specific handling for bulk_change_id
    #                         search_body["query"]["bool"]["must"].append({
    #                             "term": {f"{field}.keyword": int(value)}  # Ensure integer match
    #                         })
    #                     else:
    #                         search_body["query"]["bool"]["must"].append({
    #                             "term": {f"{field}.keyword": value}
    #                         })
    #                 else:
    #                     search_body["query"]["bool"]["must"].append({
    #                         "terms": {f"{field}.keyword": values}
    #                     })
    #             elif isinstance(values, bool):  # Single boolean condition
    #                 search_body["query"]["bool"]["must"].append({
    #                     "term": {f"{field}": values}
    #                 })
    #             elif isinstance(values, int):  # Single integer condition
    #                 search_body["query"]["bool"]["must"].append({
    #                     "term": {f"{field}": values}
    #                 })
    #             elif isinstance(values, float):  # Single integer condition
    #                 search_body["query"]["bool"]["must"].append({
    #                     "term": {f"{field}": values}
    #                 })
    #             elif isinstance(values, str):  # Single integer condition
    #                 search_body["query"]["bool"]["must"].append({
    #                     "term": {f"{field}": values}
    #                 })

    if filters:
        for field, values in filters.items():
            if values:  # Check if values is not empty
                field_type = get_field_type(index_list, field)
                print(f"Field: {field}, Field Type: {field_type}")
                # Check if field type is 'text' and add .keyword to the field name
                # field_query = f"{field}.keyword" if field_type == 'text' else field
                if isinstance(values, list):  # Handle lists
                    if len(values) == 1:
                        # Single value, handle boolean or other types
                        value = values[0]

                        if field == ('att_certified') and isinstance(value, str) and value.lower() in ['true', 'false']:
                        # Handle boolean as string
                            search_body["query"]["bool"]["must"].append({
                                "term": {f"{field}": value.lower()}  # Converts to a boolean
                            })
                        elif field_type == "text":
                            search_body["query"]["bool"]["must"].append({
                                "term": {f"{field}.keyword": value }
                            })
                        elif field in ['nsdev', 'volte_capable', 'assigned'] and isinstance(value, str) and value.lower() in ['yes', 'no']:
                            # Convert 'yes'/'no' to boolean
                            boolean_value = True if value.lower() == 'yes' else False
                            search_body["query"]["bool"]["must"].append({
                                "term": {f"{field}": boolean_value}  # Use the converted boolean value
                            })
                        elif isinstance(value, bool):  # Handle boolean condition
                            search_body["query"]["bool"]["must"].append({
                                "term": {f"{field}": value.lower()}
                            })
                        elif isinstance(value, (list, str)) and field == "billing_cycle_end_date":
                            try:
                                if isinstance(value, str):
                                    value = [value]
                                formatted_values = convert_date_format(value)
                                # print("@@@@@@@@@@", formatted_values)
                                search_body["query"]["bool"]["must"].append({
                                    "terms": {f"{field}": formatted_values}
                                })
                            except ValueError:
                                continue
                        elif isinstance(value, (list, str)) and field_type == "date":
                            if isinstance(value, str):
                                value = [value]
                            for item in value:  # Iterate through each item in the list
                                if isinstance(item, str):  # Ensure the item is a string
                                    try:
                                        parsed_date = datetime.strptime(item, "%m/%d/%Y %H:%M:%S")  # Adjust format if necessary
                                        print("$$$$$$$ ---", parsed_date)
                                        formatted_value = parsed_date.strftime("%Y-%m-%dT%H:%M:%S")
                                        print("%%%%%% ---", formatted_value)
                                        search_body["query"]["bool"]["must"].append({
                                            "term": {f"{field}": formatted_value}
                                        })
                                    except ValueError:
                                        continue
                        elif isinstance(value, (list, str)):
                            try:
                                if isinstance(value, str):
                                    value = [value]  # Convert single string to a list if it's not already a list

                                for item in value:
                                    if isinstance(item, str):  # Ensure it's a string before adding to the query
                                        # Create a match query for each item in the list
                                        search_body["query"]["bool"]["must"].append({
                                            "match": {
                                                f"{field}": item  # Using match for full-text search
                                            }
                                        })
                            except ValueError:
                                continue

                        elif isinstance(value, bool):  # Handle boolean condition
                            search_body["query"]["bool"]["must"].append({
                                "term": {f"{field}": value.lower()}
                            })
                        elif isinstance(value, str) and field == 'status'  and index_list=='sim_management_bulk_change_request':  # Handle integer values
                            search_body["query"]["bool"]["must"].append({
                                "term": {f"{field}.keyword": value}
                             })
                        elif isinstance(value, str):  # Handle integer values
                            search_body["query"]["bool"]["must"].append({
                                "term": {f"{field}": value}
                             })
                        else:
                            search_body["query"]["bool"]["must"].append({
                                "term": {f"{field}.keyword": value}
                            })
                    else:
                        if field_type == "text":
                            search_body["query"]["bool"]["must"].append({
                                "terms": {f"{field}.keyword": values }
                            })
                        elif isinstance(values, (list, str)) and field == "billing_cycle_end_date":
                            if isinstance(values, str):
                                values = [values]
                            formatted_values = []
                            for item in values:
                                if isinstance(item, str):
                                    try:
                                        formatted_value = convert_date_format(item)
                                        formatted_values.append(formatted_value)
                                    except ValueError:
                                        continue
                            if formatted_values:
                                search_body["query"]["bool"]["must"].append({
                                    "terms": {f"{field}": formatted_values}
                                })
                        elif isinstance(values, (list, str))  and field_type == "date":
                            if isinstance(values, str):
                                values = [values]
                            formatted_values = []
                            for item in values:
                                if isinstance(item, str):
                                    try:
                                        parsed_date = datetime.strptime(item, "%m/%d/%Y %H:%M:%S")
                                        formatted_value = parsed_date.strftime("%Y-%m-%dT%H:%M:%S")
                                        formatted_values.append(formatted_value)
                                    except ValueError:
                                        print(f"Invalid date format for item: {item}")
                                        continue
                            if formatted_values:
                                search_body["query"]["bool"]["must"].append({
                                    "terms": {f"{field}": formatted_values}
                                })
                        elif isinstance(values, (list, str)):
                            if isinstance(values, str):
                                values = [values]
                            formatted_values = []
                            for item in values:
                                if isinstance(item, str):  # Ensure it's a string before adding to the query
                                    try:
                                        # Create a match query for each item in the list
                                        formatted_values.append(item)
                                    except ValueError:
                                        continue
                            if formatted_values:
                                search_body["query"]["bool"]["must"].append({
                                    "terms": {
                                        f"{field}": formatted_values  # List of values for multiple exact matches
                                    }
                                })
                        else:
                            # Multiple values in the list
                            search_body["query"]["bool"]["must"].append({
                                "terms": {f"{field}": values}
                            })
                elif isinstance(values, bool):  # Single boolean condition
                    search_body["query"]["bool"]["must"].append({
                        "term": {f"{field}": values}
                    })
                elif isinstance(values, int):
                    search_body["query"]["bool"]["must"].append({
                        "term": {f"{field}": values}
                    })
                elif isinstance(values, str):
                    search_body["query"]["bool"]["must"].append({
                        "term": {f"{field}": values}
                    })


    print("Search body:", search_body)

    try:
        # if db_name=="altaworx_central":
        #     index_list=index_list
        # else:
        #     index_list=f"{index_list}_{db_name}"
        # print("index_name------",index_list)
        response = es.search(index=index_list, body=search_body)
        # print("Response:", response)  # Debugging line
        cleaned_response = replace_none_with_empty(response)
        return cleaned_response
    except Exception as e:
        print(f"An error occurred while searching: {e}")
        return None

def replace_none_with_empty(data):
    """
    Recursively replace 'None' values with empty strings in the given dictionary or list.
    """
    if isinstance(data, dict):
        return {key: replace_none_with_empty(value) for key, value in data.items()}
    elif isinstance(data, list):
        return [replace_none_with_empty(item) for item in data]
    elif data is None:
        return ""  # Replace None with an empty string
    else:
        return data

# def advanced_search_data(filters, index_list, start, end,flag,scroll_time='1m'):
#     # Initialize the search body
#     search_body = {
#         "query": {
#             "bool": {
#                 "must": []
#             }
#         },
#         "size": 10000,  # Fetch up to 10,000 records per scroll batch
#         "track_total_hits": True  # Ensure we track the total number of hits
#     }

#     if filters:
#         for field, values in filters.items():
#             if values:  # Check if values is not empty
#                 if isinstance(values, list):  # Handle lists
#                     if len(values) == 1:
#                         # Single value, handle boolean or other types
#                         value = values[0]
#                         if isinstance(value, bool):  # Handle boolean condition
#                             search_body["query"]["bool"]["must"].append({
#                                 "term": {f"{field}": value}
#                             })
#                         elif field == 'bulk_change_id':  # Specific handling for bulk_change_id
#                             search_body["query"]["bool"]["must"].append({
#                                 "term": {field: int(value)}  # Ensure integer match
#                             })
#                         else:
#                             search_body["query"]["bool"]["must"].append({
#                                 "term": {f"{field}.keyword": value}
#                             })
#                     else:
#                         # Multiple values, use terms query
#                         search_body["query"]["bool"]["must"].append({
#                             "terms": {f"{field}.keyword": values}
#                         })
#                 elif isinstance(values, bool):  # Single boolean condition
#                     search_body["query"]["bool"]["must"].append({
#                         "term": {f"{field}": values}
#                     })


#     print("Search body:", search_body)

#     try:
#         response = es.search(index=index_list, body=search_body, scroll=scroll_time)


#         scroll_id = response.get('_scroll_id')
#         if not scroll_id:
#             print("Error: No scroll ID returned.")
#             return None

#         # Retrieve the hits
#         all_hits = response['hits']['hits']
#         print(f"Fetched {len(all_hits)} initial records.")

#         # Fetch subsequent batches using the scroll API
#         while True:
#             scroll_response = es.scroll(scroll_id=scroll_id, scroll=scroll_time)
#             hits = scroll_response.get('hits', {}).get('hits', [])

#             if not hits:
#                 print("No more records found.")
#                 break

#             all_hits.extend(hits)
#             print(f"Fetched {len(hits)} more records, total: {len(all_hits)}.")

#             # Update the scroll ID
#             scroll_id = scroll_response.get('_scroll_id')
#             if not scroll_id:
#                 print("Error: No scroll ID found in subsequent response.")
#                 break

#         # Clear the scroll context
#         es.clear_scroll(scroll_id=scroll_id)
#         print(f"Total records fetched: {len(all_hits)}")

#         # Write the results to S3
#         # write_to_s3(all_hits, s3_bucket, s3_key)
#         # print(f"Results written to S3: {s3_bucket}/{s3_key}")

#         return all_hits

#     except Exception as e:
#         print(f"An error occurred while searching: {e}")
#         return None


def advanced_search_data(filters, index_list, start, end,flag,db_name,scroll_time='1m'):
    # Initialize the search body
    if db_name=="altaworx_central":
            index_list=index_list
    else:
        index_list=f"{index_list}_{db_name}"
    print("index_name------",index_list)

    search_body = {
        "query": {
            "bool": {
                "must": []
            }
        },
        "size": 10000,  # Fetch up to 10,000 records per scroll batch
        "track_total_hits": True  # Ensure we track the total number of hits
    }

    if filters:
        for field, values in filters.items():
            if values:  # Check if values is not empty
                field_type = get_field_type(index_list, field)
                print(f"Field: {field}, Field Type: {field_type}")
                # Check if field type is 'text' and add .keyword to the field name
                # field_query = f"{field}.keyword" if field_type == 'text' else field
                if isinstance(values, list):  # Handle lists
                    if len(values) == 1:
                        # Single value, handle boolean or other types
                        value = values[0]

                        if field == ('att_certified') and isinstance(value, str) and value.lower() in ['true', 'false']:
                        # Handle boolean as string
                            search_body["query"]["bool"]["must"].append({
                                "term": {f"{field}": value.lower()}  # Converts to a boolean
                            })
                        elif field_type == "text":
                            search_body["query"]["bool"]["must"].append({
                                "term": {f"{field}.keyword": value }
                            })
                        elif field in ['nsdev', 'volte_capable', 'assigned'] and isinstance(value, str) and value.lower() in ['yes', 'no']:
                            # Convert 'yes'/'no' to boolean
                            boolean_value = True if value.lower() == 'yes' else False
                            search_body["query"]["bool"]["must"].append({
                                "term": {f"{field}": boolean_value}  # Use the converted boolean value
                            })
                        elif isinstance(value, bool):  # Handle boolean condition
                            search_body["query"]["bool"]["must"].append({
                                "term": {f"{field}": value.lower()}
                            })
                        elif isinstance(value, (list, str)) and field == "billing_cycle_end_date":
                            try:
                                if isinstance(value, str):
                                    value = [value]
                                formatted_values = convert_date_format(value)
                                # print("@@@@@@@@@@", formatted_values)
                                search_body["query"]["bool"]["must"].append({
                                    "terms": {f"{field}": formatted_values}
                                })
                            except ValueError:
                                continue
                        elif isinstance(value, (list, str)) and field_type == "date":
                            if isinstance(value, str):
                                value = [value]
                            for item in value:  # Iterate through each item in the list
                                if isinstance(item, str):  # Ensure the item is a string
                                    try:
                                        parsed_date = datetime.strptime(item, "%m/%d/%Y %H:%M:%S")  # Adjust format if necessary
                                        print("$$$$$$$ ---", parsed_date)
                                        formatted_value = parsed_date.strftime("%Y-%m-%dT%H:%M:%S")
                                        print("%%%%%% ---", formatted_value)
                                        search_body["query"]["bool"]["must"].append({
                                            "term": {f"{field}": formatted_value}
                                        })
                                    except ValueError:
                                        continue
                        elif isinstance(value, (list, str)):
                            try:
                                if isinstance(value, str):
                                    value = [value]  # Convert single string to a list if it's not already a list

                                for item in value:
                                    if isinstance(item, str):  # Ensure it's a string before adding to the query
                                        # Create a match query for each item in the list
                                        search_body["query"]["bool"]["must"].append({
                                            "match": {
                                                f"{field}": item  # Using match for full-text search
                                            }
                                        })
                            except ValueError:
                                continue

                        elif isinstance(value, bool):  # Handle boolean condition
                            search_body["query"]["bool"]["must"].append({
                                "term": {f"{field}": value.lower()}
                            })
                        elif isinstance(value, str) and field == 'status'  and index_list=='sim_management_bulk_change_request':  # Handle integer values
                            search_body["query"]["bool"]["must"].append({
                                "term": {f"{field}.keyword": value}
                             })
                        elif isinstance(value, str):  # Handle integer values
                            search_body["query"]["bool"]["must"].append({
                                "term": {f"{field}": value}
                             })
                        else:
                            search_body["query"]["bool"]["must"].append({
                                "term": {f"{field}.keyword": value}
                            })
                    else:
                        if field_type == "text":
                            search_body["query"]["bool"]["must"].append({
                                "terms": {f"{field}.keyword": values }
                            })
                        elif isinstance(values, (list, str)) and field == "billing_cycle_end_date":
                            if isinstance(values, str):
                                values = [values]
                            formatted_values = []
                            for item in values:
                                if isinstance(item, str):
                                    try:
                                        formatted_value = convert_date_format(item)
                                        formatted_values.append(formatted_value)
                                    except ValueError:
                                        continue
                            if formatted_values:
                                search_body["query"]["bool"]["must"].append({
                                    "terms": {f"{field}": formatted_values}
                                })
                        elif isinstance(values, (list, str)):
                            if isinstance(values, str):
                                values = [values]
                            formatted_values = []
                            for item in values:
                                if isinstance(item, str) and field_type == "date":
                                    try:
                                        parsed_date = datetime.strptime(item, "%m/%d/%Y %H:%M:%S")
                                        formatted_value = parsed_date.strftime("%Y-%m-%dT%H:%M:%S")
                                        formatted_values.append(formatted_value)
                                    except ValueError:
                                        print(f"Invalid date format for item: {item}")
                                        continue
                            if formatted_values:
                                search_body["query"]["bool"]["must"].append({
                                    "terms": {f"{field}": formatted_values}
                                })
                        elif isinstance(values, (list, str)):
                            if isinstance(values, str):
                                values = [values]
                            formatted_values = []
                            for item in values:
                                if isinstance(item, str):  # Ensure it's a string before adding to the query
                                    try:
                                        # Create a match query for each item in the list
                                        formatted_values.append(item)
                                    except ValueError:
                                        continue
                            if formatted_values:
                                search_body["query"]["bool"]["must"].append({
                                    "terms": {f"{field}": formatted_values}
                                })
                        else:
                            # Multiple values in the list
                            search_body["query"]["bool"]["must"].append({
                                "terms": {f"{field}": values}
                            })
                elif isinstance(values, bool):  # Single boolean condition
                    search_body["query"]["bool"]["must"].append({
                        "term": {f"{field}": values}
                    })
                elif isinstance(values, int):
                    search_body["query"]["bool"]["must"].append({
                        "term": {f"{field}": values}
                    })
                elif isinstance(values, str):
                    search_body["query"]["bool"]["must"].append({
                        "term": {f"{field}": values}
                    })


    print("Search body:", search_body)
    print("******************",index_list)

    try:
        response = es.search(index=index_list, body=search_body, scroll=scroll_time)


        scroll_id = response.get('_scroll_id')
        if not scroll_id:
            print("Error: No scroll ID returned.")
            return None

        # Retrieve the hits
        all_hits = response['hits']['hits']
        print(f"Fetched {len(all_hits)} initial records.")
        total_records = response['hits']['total']['value']

        print("**********************************",total_records)

        if flag =="pagination":  # Check if flag is set for paginated results (scrolling)
                fetched_records = all_hits
                while True:
                    scroll_response = es.scroll(scroll_id=scroll_id, scroll=scroll_time)
                    hits = scroll_response.get('hits', {}).get('hits', [])

                    if not hits:
                        print("No more records found.")
                        break

                    fetched_records.extend(hits)
                    print(f"Fetched {len(hits)} more records, total: {len(fetched_records)}.")

                    # Update scroll_id for the next batch
                    scroll_id = scroll_response.get('_scroll_id')
                    if not scroll_id:
                        print("Error: No scroll ID found in subsequent response.")
                        break

                # Return the records for the pagination range requested (start to end)
                paginated_results = fetched_records[start:end]
                es.clear_scroll(scroll_id=scroll_id)  # Clear the scroll context
                return paginated_results,total_records

        else:
                # If flag is not set, just return the first batch of results
                return all_hits,total_records


    except Exception as e:
        print(f"An error occurred while searching: {e}")
        return None




#used to check through multiple index but currently not using
def search_across_indexes(query, index_list):
    data_set = ast.literal_eval(index_list)
    index_list = list(data_set)
    search_body = {
        "query": {
            "bool": {
                "should": [
                    {
                        "multi_match": {
                            "query": query,
                            "fields": ["*"],  # Fields from index1
                            "boost": 1.0,
                            "type":"most_fields"
                        }
                    },
                    {
                        "multi_match": {
                            "query": query,
                            "fields": ["*"],  # Fields from index2
                            "boost": 1.0,
                            "type":"most_fields"
                        }
                    },
                    {
                        "multi_match": {
                            "query": query,
                            "fields": ["*"],  # Fields from index3
                            "boost": 1.0,
                            "type":"most_fields"
                        }
                    }
                ]
            }
        }
    }

    try:
        response = es.search(index=index_list, body=search_body)
        return response
    except Exception as e:
        print(f"An error occurred while searching: {e}")
        return None


def convert_to_tenant_timezone(date_str, tenant_time_zone):
    """Converts date from Asia/Kolkata time zone to tenant's time zone (e.g., America/Chicago)."""
    try:
        date_obj = datetime.fromisoformat(date_str)
        if date_obj.tzinfo is None:
            date_obj = pytz.utc.localize(date_obj)
        tenant_timezone = pytz.timezone(tenant_time_zone)
        # kolkata_timezone = pytz.timezone("Asia/Kolkata")
        # date_obj = kolkata_timezone.localize(date_obj)
        date_obj = date_obj.astimezone(tenant_timezone)
        print("*****", date_obj.strftime("%m-%d-%Y %H:%M:%S"))
        return date_obj.strftime("%m-%d-%Y %H:%M:%S")
    except ValueError:
        return date_str

# AWS S3 client
s3_client = boto3.client('s3')


def export_inventory(data):
    S3_BUCKET_NAME = 'searchexcelssandbox'
    try:
        logging.info("data--------",data)
        search_all_columns = data.get("search_all_columns", "true")
        index_name=data.get('index_name')
        cols=data.get("cols")
        search=data.get("search")
        db_name=data.get("db_name","")
        bulk_change_id=data.get("bulk_change_id",None)
        flag=data.get("flag",None)
        try:
            pages=data.get("pages")
            start=int(pages.get("start",None))
            end=int(pages.get("end",10000))
            logging.info("^^^^^^^^^^",start,end)
        except:
            pass
        results=None
        sources=[]
        pquery=f"select search_tables,index_search_type from open_search_index where search_module={index_name}"
        logging.info("########",pquery)
        logging.info("dbname----",db_name)
        with psycopg2.connect(
            dbname="altaworx_test",
            user="root",
            password="AmopTeam123",
            host="amopuatpostgresoct23.c3qae66ke1lg.us-east-1.rds.amazonaws.com",
            port="5432"
        ) as conn:
            with conn.cursor() as cur:
                pquery = f"SELECT search_tables, index_search_type FROM open_search_index WHERE search_module='{index_name}'"
                cur.execute(pquery)
                data1 = cur.fetchall()
        logging.info("---------data1",data1)
        index_list=data1[0][0]
        search_type=data1[0][1]
        #search_type="whole"

        tenant_name = data.get("partner", "")  # Assuming 'partner' is the correct key for tenant name
        if not tenant_name:
            raise ValueError("Tenant name is missing.")
        tenant_name = "Altaworx Test"
        tenant_timezone_query = f"""SELECT time_zone FROM tenant WHERE tenant_name = %s"""
        logging.info("@@@@", tenant_timezone_query)

        with psycopg2.connect(
            dbname="common_utils",
            user="root",
            password="AmopTeam123",
            host="amopuatpostgresoct23.c3qae66ke1lg.us-east-1.rds.amazonaws.com",
            port="5432"
        ) as conn:
            with conn.cursor() as cur:
                cur.execute(tenant_timezone_query, (tenant_name,))
                tenant_timezone = cur.fetchone()
                logging.info(f"Retrieved tenant time zone: {tenant_timezone}")  # Debug print

        if tenant_timezone is None or tenant_timezone[0] is None:
            logging.info(f"No valid timezone found for tenant '{tenant_name}'. Using UTC as fallback.")
            tenant_time_zone = "UTC"  # Use UTC as default
        else:
            tenant_time_zone = tenant_timezone[0]

        logging.info(f"Using time zone: {tenant_time_zone}")  # Debug print

        match = re.search(r"\(\w+\s[+\-]?\d{2}:\d{2}:\d{2}\)\s*(Asia\s*/\s*Kolkata)", tenant_time_zone)
        if match:
            tenant_time_zone = "Asia/Kolkata"


        if search=="advanced":
            filters = data.get('filters')
            # Add tenant_id and is_active as string fields
            if is_field_present(index_list, 'is_active'):
                filters['is_active'] = True
            if is_field_present(index_list, 'tenant_id'):
                tenant_id=1
                if index_name == "sim_management_inventory" or index_name == "vw_combined_device_inventory_export":
                    filters['tenant_id'] = tenant_id
                else:
                    filters.pop('tenant_id', None)
            if 'bulk_change_id' in filters:
                bulk_change_id=filters['bulk_change_id']
                if index_name != 'sim_management_bulk_change_request':
                    removed_value = filters.pop('bulk_change_id', None)
                else:
                    # bulk_change_id=int(bulk_change_id)
                    filters['bulk_change_id'] = bulk_change_id
            logging.info("filters are ----", filters)
            flag="pagination"
            if start > 9999 :
                search="advanced_pro"
                results,total_rows =advanced_search_data(filters, index_list, start, end,flag,db_name,scroll_time='1m')
            else:
                results=advance_search_data(filters,index_list,start, end,db_name)

        elif search=="whole":
            query=data.get("search_word")
            module_name=get_module_name(index_name,db_name)
            columns=get_columns_from_db(module_name)
            logging.info("Column",columns)
            if index_list == "sim_management_inventory_action_history":
                custom_filter={}
            elif index_list == "sim_management_inventory" or index_name == "vw_combined_device_inventory_export":
                custom_filter={"is_active": True, "tenant_id": 1 }
            elif flag == "bulk_history" or index_list == "sim_management_bulk_change_request":
                custom_filter={"bulk_change_id": bulk_change_id[0] }
                columns.append("bulk_change_id")
            else:
                custom_filter={"is_active": True}
            logging.info("custom_filter------------",custom_filter)
            results = search_data(query, search_all_columns,index_list,start,end,columns,db_name,custom_filter)
            logging.info("Result_______________",results)
        else:
            logging.info("incorrect type")
        total_rows=0
        if results:
            if search =="whole" or search == "advanced":
                hits = results.get('hits', {}).get('hits', [])
                total_rows = results['hits']['total']['value']
            if search == "advanced_pro":
                hits= results
                total_rows = total_rows
            for hit in hits:
                source = hit.get('_source', {})
                for k, v in source.items():
                    if isinstance(v, str) and 'T' in v:  # If the value looks like a date
                        source[k] = convert_to_tenant_timezone(v, tenant_time_zone)
                source = {k: str(v) for k, v in source.items()}
                sources.append(source)


            # Filter the columns if 'cols' are provided
            if cols:
                sources = [{key: value for key, value in source.items() if key in cols} for source in sources]
            #return {"flag":True,"results":sources}
            logging.info("*************",sources)
            # Create a DataFrame for the filtered data
            df = pd.DataFrame(sources)
            required_columns = [
                "service_provider_display_name", "imei", "carrier_rate_plan_name", "carrier_cycle_usage_mb",
                "sms_count", "cost_center", "account_number", "customer_name", "date_added", "iccid",
                "msisdn", "eid", "username", "carrier_cycle_usage_bytes",
                "customer_rate_pool_name", "customer_rate_plan_name", "sim_status",
                "date_activated", "ip_address", "carrier_rate_plan_display_rate", "communication_plan",
                "pool_id", "telegence_feature", "plan_limit_mb",
                "customer_data_allocation_mb", "service_zip_code", "device_make", "device_model",
                "next_bill_cycle_date", "billing_account_number", "foundation_account_number",
                "ban_status", "minutes_used", "modified_by", "modified_date"
            ]
            # Filter the DataFrame to include only the required columns
            df = df[required_columns]

            # Rename the columns to match the SQL aliases
            df.rename(columns={
                "service_provider_display_name": "service_provider_name",
                "carrier_rate_plan_name": "rate_plan",
                "carrier_cycle_usage_mb": "data_usage_mb",
                "carrier_cycle_usage_bytes": "data_usage_bytes",
                "sim_status": "status_display_name",
                "carrier_rate_plan_display_rate": "carrier_rate_plan_cost",
                "plan_limit_mb": "carrier_data_allocation_mb",
                "telegence_feature": "feature_codes"
            }, inplace=True)
            # Convert all large integers to strings to avoid scientific notation
            # Ensure all numeric columns are converted to plain strings
            for col in df.columns:
                if df[col].dtype in ['int64', 'float64']:
                    df[col] = df[col].apply(lambda x: str(int(x)) if pd.notnull(x) else '')
            # Filter the DataFrame to include only the required columns
            #df = df[required_columns]
            # df.columns = [col.replace('_', ' ').capitalize() for col in df.columns]
            # df.columns = [" ".join(word.capitalize() for word in col.split("_")) for col in df.columns]
            acronyms = {"SMS", "IMEI", "IP", "BAN", "URL", "UID", "MAC", "EID", "MSISDN", "MB", "CCID", "ICCID", "MSISDN", "IP"}

            # Define your special cases for capitalization (e.g., "And" -> "and" and "Att" -> "AT&T")
            special_replacements = {
                "Att": "AT&T",
                "And": "and"
            }
            # Format column names
            df.columns = [
                " ".join(
                    part.upper() if part.upper() in acronyms else part.capitalize()
                    for part in col.split("_")
                )
                for col in df.columns
            ]

            # Apply special replacements (Att -> AT&T, And -> and)
            df.columns = [
                " ".join([special_replacements.get(word, word) for word in col.split(" ")]) for col in df.columns
            ]




            df['S.No'] = range(1, len(df) + 1)
            columns = ['S.No'] + [col for col in df.columns if col != 'S.No']
            df = df[columns]
            # # Convert to CSV (you can convert to Excel if you prefer)
            # csv_buffer = StringIO()
            # df.to_csv(csv_buffer, index=False)

            # # Upload the CSV file to S3
            # file_name = f"exports/search/Inventory_export.csv"
            # csv_buffer.seek(0)  # Move to the start of the StringIO buffer

            # # Upload to S3 (public or private based on your needs)
            # s3_client.put_object(
            #     Bucket=S3_BUCKET_NAME,
            #     Key=file_name,
            #     Body=csv_buffer.getvalue(),
            #     ContentType='text/csv'
            # )
            # Convert DataFrame to Excel using openpyxl
            excel_buffer = BytesIO()
            with pd.ExcelWriter(excel_buffer, engine='openpyxl') as writer:
                df.to_excel(writer, index=False, sheet_name="Inventory")
                workbook = writer.book
                sheet = writer.sheets["Inventory"]

                # Write headers with formatting
                for c_idx, column in enumerate(df.columns, start=1):
                    cell = sheet.cell(row=1, column=c_idx, value=column)
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                    cell.fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")

                # Write data rows
                for r_idx, row in enumerate(df.itertuples(index=False), start=2):
                    for c_idx, value in enumerate(row, start=1):
                        sheet.cell(row=r_idx, column=c_idx, value=value)

                # Adjust column widths
                for col_idx, col_cells in enumerate(sheet.columns, 1):
                    max_length = max(len(str(cell.value or "")) for cell in col_cells)
                    sheet.column_dimensions[get_column_letter(col_idx)].width = max_length + 2

            excel_buffer.seek(0)

            # Define the file name for the Excel file
            file_name = f"exports/search/Inventory Export.xlsx"

            # Upload the Excel file to S3
            s3_client.put_object(
                Bucket=S3_BUCKET_NAME,
                Key=file_name,
                Body=excel_buffer.getvalue(),
                ContentType='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )

            # Generate URL (public URL or pre-signed URL)
            download_url = f"https://{S3_BUCKET_NAME}.s3.amazonaws.com/{file_name}"
            search_result = {
                'flag': True,
                'download_url': download_url  # Return the URL where the file is stored in S3
            }
            return search_result
        else:
            search_result = {
                "flag": True,
                "data": {
                    "table": sources
                },
                "pages": {
                    "start": start,
                    "end": end,
                    "total": total_rows
                }
            }

        return search_result
    except Exception as e:
        print(f"Exception is {e}")
        search_result = {
                "flag": False,
                "message": "No Records found"
            }
        return search_result

def search_export(data):
    S3_BUCKET_NAME = "searchexcelssandbox"
    module_name = data.get("module_name", "")
    module_name_snake_case = "_".join(module_name.strip().lower().split())
    try:
        print("data--------",data)
        search_all_columns = data.get("search_all_columns", "true")
        index_name=data.get('index_name')
        cols=data.get("cols")
        search=data.get("search")
        db_name=data.get("db_name","altaworx_test")
        bulk_change_id=data.get("bulk_change_id",None)
        flag=data.get("flag",None)
        try:
            pages=data.get("pages")
            start=int(pages.get("start",None))
            end=int(pages.get("ending",10000))
            print("^^^^^^^^^^",start,end)
        except:
            pass
        results=None
        sources=[]
        pquery=f"select search_tables,index_search_type from open_search_index where search_module={index_name}"
        print("########",pquery)
        print("dbname----",db_name)
        with psycopg2.connect(
            dbname="altaworx_test",
            user="root",
            password="AmopTeam123",
            host="amopuatpostgresoct23.c3qae66ke1lg.us-east-1.rds.amazonaws.com",
            port="5432"
        ) as conn:
            with conn.cursor() as cur:
                pquery = f"SELECT search_tables, index_search_type FROM open_search_index WHERE search_module='{index_name}'"
                cur.execute(pquery)
                data1 = cur.fetchall()
        print("---------data1",data1)
        index_list=data1[0][0]
        search_type=data1[0][1]
        #search_type="whole"

        tenant_name = data.get("partner", "")  # Assuming 'partner' is the correct key for tenant name
        if not tenant_name:
            raise ValueError("Tenant name is missing.")
        tenant_name = "Altaworx Test"
        tenant_timezone_query = f"""SELECT time_zone FROM tenant WHERE tenant_name = %s"""
        logging.info("@@@@", tenant_timezone_query)

        with psycopg2.connect(
            dbname="common_utils",
            user="root",
            password="AmopTeam123",
            host="amopuatpostgresoct23.c3qae66ke1lg.us-east-1.rds.amazonaws.com",
            port="5432"
        ) as conn:
            with conn.cursor() as cur:
                cur.execute(tenant_timezone_query, (tenant_name,))
                tenant_timezone = cur.fetchone()
                logging.info(f"Retrieved tenant time zone: {tenant_timezone}")  # Debug print

        if tenant_timezone is None or tenant_timezone[0] is None:
            logging.info(f"No valid timezone found for tenant '{tenant_name}'. Using UTC as fallback.")
            tenant_time_zone = "UTC"  # Use UTC as default
        else:
            tenant_time_zone = tenant_timezone[0]

        logging.info(f"Using time zone: {tenant_time_zone}")  # Debug print

        match = re.search(r"\(\w+\s[+\-]?\d{2}:\d{2}:\d{2}\)\s*(Asia\s*/\s*Kolkata)", tenant_time_zone)
        if match:
            tenant_time_zone = "Asia/Kolkata"


        if search=="advanced":
            filters = data.get('filters')
            # Add tenant_id and is_active as string fields
            if is_field_present(index_list, 'is_active'):
                filters['is_active'] = True
            if is_field_present(index_list, 'tenant_id'):
                tenant_id=1
                if index_name == "sim_management_inventory" or index_name == "vw_combined_device_inventory_export":
                    filters['tenant_id'] = tenant_id
                else:
                    filters.pop('tenant_id', None)
            if 'bulk_change_id' in filters:
                bulk_change_id=filters['bulk_change_id']
                if index_name != 'sim_management_bulk_change_request':
                    removed_value = filters.pop('bulk_change_id', None)
                else:
                    # bulk_change_id=int(bulk_change_id)
                    filters['bulk_change_id'] = bulk_change_id
            logging.info("filters are ----", filters)
            flag="pagination"
            if start > 9999 :
                search="advanced_pro"
                results,total_rows =advanced_search_data(filters, index_list, start, end,flag,db_name,scroll_time='1m')
            else:
                results=advance_search_data(filters,index_list,start, end,db_name)

        elif search=="whole":
            query=data.get("search_word")
            module_name=get_module_name(index_name,db_name)
            columns=get_columns_from_db(module_name)
            logging.info("Column",columns)
            if index_list == "sim_management_inventory_action_history":
                custom_filter={}
            elif index_list == "sim_management_inventory" or index_name == "vw_combined_device_inventory_export":
                custom_filter={"is_active": True, "tenant_id": 1 }
            elif flag == "bulk_history" or index_list == "sim_management_bulk_change_request":
                custom_filter={"bulk_change_id": bulk_change_id[0] }
                columns.append("bulk_change_id")
            else:
                custom_filter={"is_active": True}
            logging.info("custom_filter------------",custom_filter)
            results = search_data(query, search_all_columns,index_list,start,end,columns,db_name,custom_filter)
            logging.info("Result_______________",results)
        else:
            logging.info("incorrect type")
        total_rows=0
        if results:
            if search =="whole" or search == "advanced":
                hits = results.get('hits', {}).get('hits', [])
                total_rows = results['hits']['total']['value']
            if search == "advanced_pro":
                hits= results
                total_rows = total_rows
            for hit in hits:
                source = hit.get('_source', {})
                for k, v in source.items():
                    if isinstance(v, str) and 'T' in v:  # If the value looks like a date
                        source[k] = convert_to_tenant_timezone(v, tenant_time_zone)
                source = {k: str(v) for k, v in source.items()}
                sources.append(source)
            print("*************",sources)
            # print("*************",sources)
            #return {"flag":True,"results":sources}
            # Filter the columns if 'cols' are provided
            if cols:
                sources = [
                    {key: value for key, value in source.items() if key in cols}
                    for source in sources
                ]

            # Create a DataFrame for the filtered data
            df = pd.DataFrame(sources)
            # required_columns = [
            #     "service_provider_display_name", "imei", "rate_plan", "data_usage_mb", "sms_count",
            #     "cost_center", "account_number", "customer_name", "date_added", "iccid",
            #     "msisdn", "eid", "username", "data_usage_bytes",
            #     "customer_rate_pool_name", "customer_rate_plan_name", "status_display_name",
            #     "date_activated", "ip_address", "billing_account_number", "foundation_account_number",
            #     "modified_by", "modified_date"
            # ]
            # Specify the required columns from the original data
            # Specify the required columns from the original data (duplicates removed)

            ##databse connenction
            database = DB("common_utils", **db_config)
            # Fetch the query from the database based on the module name
            module_query_df = database.get_data(
                "export_columns", {"module_name": module_name}
            )
            # logging.info(module_query_df,'module_query_df')
            ##checking the dataframe is empty or not
            if module_query_df.empty:
                return {
                    "flag": False,
                    "message": f"No query found for module name: {module_name}",
                }
           # Extract the required columns from the query result
            required_columns = module_query_df.iloc[0].get("required_columns")
            if not required_columns:
                print(f"Unknown module name: {module_name}")
                return {"flag": False, "message": f"Unknown module name: {module_name}"}

            # Parse the required columns as a list (assuming they are stored as a comma-separated string)
            required_columns = [col.strip().strip('"') for col in required_columns.split(",")]

            # Ensure all required columns exist in the DataFrame
            df_columns = [col.strip().strip('"') for col in df.columns]  # Normalize column names in DataFrame
            missing_columns = [col for col in required_columns if col not in df_columns]
            if missing_columns:
                return {
                    "flag": False,
                    "message": f"Missing required columns in data: {', '.join(missing_columns)}",
                }

            # Filter the DataFrame to include only the required columns
            df = df[required_columns]

            # Rename the columns to match the SQL aliases
            if module_name == "SimManagement Inventory":
                df.rename(
                    columns={
                        "service_provider_display_name": "service_provider_name",
                        "data_usage_bytes": "carrier_cycle_usage",
                        "status_display_name": "sim_status",
                        "carrier_rate_plan_display_rate": "carrier_rate_plan_cost",
                        "plan_limit_mb": "carrier_data_allocation_mb",
                        "telegence_feature": "feature_codes",
                    },
                    inplace=True,
                )

            # Filter the DataFrame to include only the required columns
            # df = df[required_columns]
            df.columns = [col.replace("_", " ").capitalize() for col in df.columns]
            df["S.No"] = range(1, len(df) + 1)
            columns = ["S.No"] + [col for col in df.columns if col != "S.No"]
            df = df[columns]
            # Create Excel Workbook
            workbook = Workbook()
            sheet = workbook.active
            sheet.title = module_name

            # Write headers
            for c_idx, column in enumerate(df.columns, start=1):
                cell = sheet.cell(row=1, column=c_idx, value=column)  # Header row (row=1)
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")

            # Write data rows
            for r_idx, row in enumerate(df.itertuples(index=False), start=2):  # Start from the second row for data
                for c_idx, value in enumerate(row, start=1):
                    sheet.cell(row=r_idx, column=c_idx, value=value)

            # Adjust column widths
            for col_idx, col_cells in enumerate(sheet.columns, 1):
                max_length = max(len(str(cell.value or "")) for cell in col_cells)
                sheet.column_dimensions[get_column_letter(col_idx)].width = max_length + 2

            # Save to BytesIO and upload to S3
            excel_buffer = BytesIO()
            workbook.save(excel_buffer)
            file_name = f"exports/search/{module_name_snake_case}.xlsx"
            excel_buffer.seek(0)

            # Upload to S3 (public or private based on your needs)
            s3_client.put_object(
                Bucket=S3_BUCKET_NAME,
                Key=file_name,
                Body=excel_buffer.getvalue(),
                ContentType="text/csv",
            )

            # Generate URL (public URL or pre-signed URL)
            download_url = f"https://{S3_BUCKET_NAME}.s3.amazonaws.com/{file_name}"
            search_result = {
                "flag": True,
                "download_url": download_url,  # Return the URL where the file is stored in S3
            }
            return search_result
        else:
            search_result = {
                "flag": True,
                "data": {"table": sources},
                "pages": {"start": start, "end": end, "total": total_rows},
            }

        return search_result
    except Exception as e:
        print(f"Exception is {e}")
        search_result = {"flag": False, "message": "No Records found"}
        return search_result

def connect_to_postgres(db_name):
    return psycopg2.connect(
        dbname=db_name,
        user="root",
        password="AmopTeam123",
        host="amoppostoct19.c3qae66ke1lg.us-east-1.rds.amazonaws.com",
        port="5432"
    )

def get_module_name(search_module,db_name):
    database = DB(db_name, **db_config)
    query = f"""
        SELECT module_name
        FROM open_search_index
        WHERE search_module = '{search_module}'
    """
    # Execute the query

    df = database.execute_query(query, True)

    if not df.empty:
        module_name = df.iloc[0]["module_name"]  # Extract the first value of the "module_name" column
        return module_name
    else:
        print("No results found.")
        return None


def get_columns_from_db( module_name):
    database = DB("common_utils", **db_config)
    query = f"""
        SELECT db_column_name
        FROM field_column_mapping
        WHERE module_name = '{module_name}' AND table_col = 'yes'
    """
    # Execute the query

    df = database.execute_query(query, True)
    # Convert the column to a list
    return df['db_column_name'].tolist()


def perform_search(data):
    print("data--------",data)
    search_all_columns = data.get("search_all_columns", "true")
    index_name=data.get('index_name')
    cols=data.get("cols")
    search=data.get("search")
    db_name=data.get("db_name")
    bulk_change_id=data.get("bulk_change_id",None)
    flag=data.get("flag",None)
    try:
        pages=data.get("pages")
        start=int(pages.get("start",None))
        end=int(pages.get("end",None))
        print("^^^^^^^^^^-----------",start,end)
    except:
        pass
    results=None
    sources=[]
    pquery=f"select search_tables,index_search_type from open_search_index where search_module={index_name}"
    print("########",pquery)
    print("dbname----",db_name)
    with psycopg2.connect(
        dbname=db_name,
        user="root",
        password="AmopTeam123",
        host="amoppostoct19.c3qae66ke1lg.us-east-1.rds.amazonaws.com",
        port="5432"
    ) as conn:
        with conn.cursor() as cur:
            pquery = f"SELECT search_tables, search_cols FROM open_search_index WHERE search_module='{index_name}'"
            cur.execute(pquery)
            data1 = cur.fetchall()
    print("---------data1",data1)
    index_list=data1[0][0]
    search_type=data1[0][1]
    #search_type="whole"


    if search=="advanced":
        filters = data.get('filters')
        # Add tenant_id and is_active as string fields
        # if is_field_present(index_list, 'is_active'):
        #     filters['is_active'] = True
        if is_field_present(index_list, 'is_active'):
            if index_list == "customerrateplan":
                toggle = data.get("toggle", None)
                is_active_value = filters.get("is_active", None)

                if not toggle:
                    # If 'is_active' is in the filters, keep it as is (true/false)
                    if is_active_value:
                        filters['is_active'] = is_active_value
                    else:
                        # If 'is_active' is not in the filters, default to 'is_active = true'
                        filters['is_active'] = ["false"]
                else:  # If toggle is True
                    filters['is_active'] = ["true"]
            else:
                # For all other indexes, ensure 'is_active' is True
                filters['is_active'] = ["true"]
        if is_field_present(index_list, 'tenant_id'):
            tenant_id=1
            if index_name == "sim_management_inventory":
                filters['tenant_id'] = tenant_id
            else:
                filters.pop('tenant_id', None)
        if 'bulk_change_id' in filters:
            bulk_change_id=filters['bulk_change_id']
            if index_name != 'sim_management_bulk_change_request':
                removed_value = filters.pop('bulk_change_id', None)
            else:
                # bulk_change_id=int(bulk_change_id)
                filters['bulk_change_id'] = bulk_change_id
        print("filters are ----", filters)
        flag="pagination"
        if start > 9999 :
            search="advanced_pro"
            results,total_rows =advanced_search_data(filters, index_list, start, end,flag,db_name,scroll_time='1m')
        else:
            results=advance_search_data(filters,index_list,start, end,db_name)

    elif search=="whole":
        query=data.get("search_word")
        module_name=get_module_name(index_name,db_name)
        columns=get_columns_from_db(module_name)
        print("Column",columns)
        if index_list == "sim_management_inventory_action_history":
            custom_filter={}
        elif index_list == "sim_management_inventory":
            custom_filter={"is_active": True, "tenant_id": 1 }
        elif flag == "bulk_history" or index_list == "sim_management_bulk_change_request":
            custom_filter={"bulk_change_id": bulk_change_id[0] }
            columns.append("bulk_change_id")
        else:
            custom_filter={"is_active": True}
        print("custom_filter------------",custom_filter)
        results = search_data(query, search_all_columns,index_list,start,end,columns,db_name,custom_filter)
        print("Result_______________",results)
    else:
        print("incorrect type")
    total_rows=0
    if results:
        if search =="whole" or search == "advanced":
            hits = results.get('hits', {}).get('hits', [])
            total_rows = results['hits']['total']['value']
        if search == "advanced_pro":
            hits= results
            total_rows = total_rows

        for hit in hits:
            source = hit.get('_source', {})
            source = {k: str(v) for k, v in source.items()}
            sources.append(source)

        #total_rows = len(sources)
        search_result = {
            "flag": True,
            "data": {
                "table": sources
            },
            "pages":{
                "start":start,
                "end":end,
                # "end": min(end, total_rows),
                "total": total_rows
            }
        }

    else:
        search_result = {
            "flag": True,
            "data": {
                "table": sources
            },
            "pages":{
                "start":start,
                "end":end,
                # "end": total_rows,
                "total": total_rows
            }
        }
    return search_result


def fetch_dropdown_old(data):
    try:
        column_name=data.get("drop_down",None)
        table=data.get("table",None)
        flag=data.get("flag",None)
        db_name=data.get("db_name",None)

        conn = psycopg2.connect(
            dbname=db_name,
            user="root",
            password="AmopTeam123",
            host="amoppostoct19.c3qae66ke1lg.us-east-1.rds.amazonaws.com",
            port="5432"
        )
        cursor = conn.cursor()
        print("-------------",column_name)

        table_query= f"SELECT search_tables, index_search_type FROM open_search_index WHERE search_module='{table}'"
        cursor.execute(table_query)
        table_data=cursor.fetchall()
        table_name=table_data[0][0]

        if  flag == "status_history":
            iccid=data.get("iccid",None)
            iccid=iccid[0]
            query = f"SELECT DISTINCT {column_name} FROM {table_name} where iccid='{iccid}' and  {column_name} IS NOT NULL"
        elif flag == "bulk_history":
            bulk_change_id=data.get("bulk_change_id",None)
            bulk_change_id=bulk_change_id
            query = f"SELECT DISTINCT {column_name} FROM {table_name} where bulk_change_id='{bulk_change_id}' and {column_name} IS NOT NULL"

        else:
            query = f"SELECT DISTINCT {column_name} FROM {table_name} WHERE {column_name} IS NOT NULL"

        # Execute the query
        cursor.execute(query)
        # Fetch all unique values and return them as a list
        unique_values = cursor.fetchall()
        col_data = []
        for value in unique_values:
            val = value[0]
            if isinstance(val, Decimal):
                col_data.append(float(val))  # Convert Decimal to float
            elif isinstance(val, bool):
                col_data.append(str(val).lower())  # Convert boolean to "true"/"false"
            elif isinstance(val, datetime):
                if column_name == "billing_cycle_end_date":
                    formatted_date = (val - pd.Timedelta(days=1)).replace(hour=23, minute=59, second=59) \
                        if val.time() == pd.Timestamp("00:00:00").time() else val
                    col_data.append(formatted_date.strftime("%m/%d/%Y %H:%M:%S"))
                else:
                    col_data.append(val.strftime("%m/%d/%Y %H:%M:%S"))
            else:
                col_data.append(val)
        # col_data = [float(value[0]) if isinstance(value[0], Decimal) else value[0] for value in unique_values]
        # col_data=[value[0] for value in unique_values]
        return {"flag": "true" , column_name:col_data}

    except Exception as e:
        print(f"An error occurred: {e}")
    finally:
        # Close the cursor and connection
        cursor.close()
        conn.close()


def fetch_dropdown(data):
    try:
        column_name = data.get("drop_down", None)
        table = data.get("table", None)
        flag = data.get("flag", None)
        db_name=data.get("db_name",None)
        cascade = data.get("cascade", {})
        toggle = data.get("toggle", None)
        print("*************",data)

        # Connect to PostgreSQL
        conn = psycopg2.connect(
            dbname=db_name,
            user="root",
            password="AmopTeam123",
            host="amoppostoct19.c3qae66ke1lg.us-east-1.rds.amazonaws.com",
            port="5432"
        )
        cursor = conn.cursor()


        # Get table name and index search type
        table_query = f"SELECT search_tables, index_search_type FROM open_search_index WHERE search_module='{table}'"
        cursor.execute(table_query)
        table_data = cursor.fetchall()
        table_name = table_data[0][0]
        tenant_id=1

        print("table_name",table_name)

        cursor.execute(f"""
            SELECT column_name
            FROM information_schema.columns
            WHERE table_name = '{table_name}' AND column_name = 'is_active'
        """)
        is_active_exists = cursor.fetchone() is not None

        cursor.execute(f"""
            SELECT column_name, data_type
            FROM information_schema.columns
            WHERE table_name = '{table_name}' AND column_name = '{column_name}';
        """)
        data_type = cursor.fetchone()
        data_type= data_type[0]
        print(data_type)

        cursor.execute(f"""
            SELECT column_name
            FROM information_schema.columns
            WHERE table_name = '{table_name}' AND column_name = 'tenant_id'
        """)
        tenant_id_exists = cursor.fetchone() is not None

        where_conditions=[]

        if is_active_exists:
            where_conditions.append("is_active = True")

        # Start building the query
        if table_name == "sim_management_inventory":
            if tenant_id_exists:
                where_conditions.append(f"tenant_id = {tenant_id}")

        if table_name == "customerrateplan":
            if not toggle:
                if "is_active = True" in where_conditions:
                    where_conditions.remove("is_active = True")

        # Handle different flags for custom queries
        if flag == "status_history":
            iccid = data.get("iccid", None)
            iccid = iccid[0] if iccid else None
            where_conditions.append(f"iccid='{iccid}'")
        elif flag == "bulk_history":
            bulk_change_id = data.get("bulk_change_id", None)
            where_conditions.append(f"bulk_change_id='{bulk_change_id}'")

        # Handle the cascade filtering
        for field, values in cascade.items():
            # Skip if the list of values is empty
            if not values:
                continue
            # If column_name matches the cascade key, skip the filter for this column
            if field == column_name:
                continue
            if values:
                # Add condition for each cascade field and its values
                where_conditions.append(f"{field} IN ({', '.join([repr(v) for v in values])})")

        print("############",where_conditions)

        # Construct the final query based on the conditions
        query = f"SELECT DISTINCT {column_name} FROM {table_name} WHERE {column_name} IS NOT NULL"
        if where_conditions:
            query += " AND " + " AND ".join(where_conditions)

        condition = 'ASC'
        if data_type in ['timestamp without time zone']:
            condition = 'DESC'


        temp = f'ORDER BY {column_name} {condition}'
        query +=" "+temp

        print("Query",query)

        # Execute the query
        cursor.execute(query)

        # Fetch all unique values and return them as a list
        unique_values = cursor.fetchall()
        col_data = []
        for value in unique_values:
            if isinstance(value[0], str):
                val = value[0].strip()  # Remove leading and trailing whitespace
            else:
                val = value[0]  # Use the value as is for non-string types

            if val == '':  # Skip empty or blank values
                continue
            elif isinstance(val, Decimal):
                col_data.append(float(val))  # Convert Decimal to float
            elif isinstance(val, bool):
                # Check if the column name is `nsdev` or `volte_capable`
                if column_name in ["nsdev", "volte_capable", "assigned"]:
                    col_data.append("Yes" if val else "No")  # Show "Yes" or "No" for these columns
                # elif column_name == "is_active" and table_name == "customerrateplan":
                #     col_data.append("Active" if val else "Inactive")
                else:
                    col_data.append(str(val).lower())
            elif isinstance(val, bool):
                col_data.append(str(val).lower())  # Convert boolean to "true"/"false"
            elif isinstance(val, datetime):
                if column_name == "billing_cycle_end_date":
                    formatted_date = (val - pd.Timedelta(days=1)).replace(hour=23, minute=59, second=59) \
                        if val.time() == pd.Timestamp("00:00:00").time() else val
                    col_data.append(formatted_date.strftime("%m/%d/%Y %H:%M:%S"))
                else:
                    col_data.append(val.strftime("%m/%d/%Y %H:%M:%S"))
            else:
                col_data.append(val)


        if column_name == 'device_count':
            col_data = [int(i) for i in col_data]
            ranges = get_ranges(col_data)
            #print(ranges)
            col_data = ranges

        return {"flag": "true", column_name: col_data}


    except Exception as e:
        print(f"An error occurred: {e}")
    finally:
        # Close the cursor and connection
        cursor.close()
        conn.close()
