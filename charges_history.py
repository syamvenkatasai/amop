"""
@Author1 : Phaneendra.Y
Created Date: 11-11-24
"""

# Importing the necessary Libraries
import os
import time
import json
import io
import re
import base64
import boto3
import zipfile
import datetime
from datetime import datetime, timedelta
from io import BytesIO
from pytz import timezone


# Third-party imports
import pandas as pd


# Internal imports
from common_utils.db_utils import DB
from common_utils.logging_utils import Logging

from openpyxl.styles import Alignment, PatternFill
from openpyxl.utils import get_column_letter


db_config = {
    "host": os.environ["HOST"],
    "port": os.environ["PORT"],
    "user": os.environ["USER"],
    "password": os.environ["PASSWORD"],
}

# AWS S3 client
s3_client = boto3.client("s3")
# Retrieve the S3 bucket name from an environment variable
S3_BUCKET_NAME = os.getenv("S3_BUCKET_NAME")

logging = Logging(name="charges_history")


def path_fun(path,data):

    if path == "/get_charge_history_data":
        result = get_charge_history_data(data)
    elif path == "/export_row_data_customer_charges":
        result = export_row_data_customer_charges(data)
    elif path == "/customers_sessions_customer_charges_export_dropdown_data":
        result = customers_sessions_customer_charges_export_dropdown_data(data)
    elif path == "/export_customer_charges":
        result = export_customer_charges(data)
    elif path == "/get_export_status":
        result = get_export_status(data)

    else:
        result = {"error": "Invalid path or method"}
        logging.warning("Invalid path or method requested: %s", path)

    return result

def get_headers_mapping(
    tenant_database,
    module_list,
    role,
    user,
    tenant_id,
    sub_parent_module,
    parent_module,
    data,
    common_utils_database,
):
    """
    Description: The  function retrieves and organizes field mappings,headers,and module features
    based on the provided module_list, role, user, and other parameters.
    It connects to a database, fetches relevant data, categorizes fields,and
    compiles features into a structured dictionary for each module.
    """
    logging.info(
        f"""
    tenant_database :{tenant_database} , user : {user} ,parent_module :{parent_module} and
    sub_parent_module:{sub_parent_module}
    """
    )
    feature_module_name = data.get("feature_module_name", "")
    user_name = data.get("username") or data.get("user_name") or data.get("user")
    tenant_name = data.get("tenant_name") or data.get("tenant")
    try:
        tenant_id = common_utils_database.get_data(
            "tenant", {"tenant_name": tenant_name}, ["id"]
        )["id"].to_list()[0]
    except Exception as e:
        logging.exception(f"Getting exception at fetching tenant id {e}")
    ret_out = {}
    # Iterate over each module name in the provided module list
    for module_name in module_list:
        # logging.debug("Processing module: %s", module_name)
        out = common_utils_database.get_data(
            "field_column_mapping", {"module_name": module_name}
        ).to_dict(orient="records")
        pop_up = []
        general_fields = []
        table_fileds = {}
        # Categorize the fetched data based on field types
        for data in out:
            if data["pop_col"]:
                pop_up.append(data)
            elif data["table_col"]:
                table_fileds.update(
                    {
                        data["db_column_name"]: [
                            data["display_name"],
                            data["table_header_order"],
                        ]
                    }
                )
            else:
                general_fields.append(data)
        # Create a dictionary to store categorized fields
        headers = {}
        headers["general_fields"] = general_fields
        headers["pop_up"] = pop_up
        headers["header_map"] = table_fileds
        try:
            final_features = []

            # Fetch all features for the 'super admin' role
            if role.lower() == "super admin":
                all_features = common_utils_database.get_data(
                    "module_features", {"module": feature_module_name}, ["features"]
                )["features"].to_list()
                if all_features:
                    final_features = json.loads(all_features[0])
            else:
                final_features = get_features_by_feature_name(
                    user_name, tenant_id, feature_module_name, common_utils_database
                )
                logging.info(
                    "Fetched features for user '%s': %s", user_name, final_features
                )

        except Exception as e:
            logging.info(f"there is some error {e}")
            pass
        # Add the final features to the headers dictionary
        headers["module_features"] = final_features
        ret_out[module_name] = headers
    # logging.info("Completed headers mapping retrieval.")
    return ret_out


def get_features_by_feature_name(
    user_name, tenant_id, feature_name, common_utils_database
):
    # Fetch user features from the database
    user_features_raw = common_utils_database.get_data(
        "user_module_tenant_mapping",
        {"user_name": user_name, "tenant_id": tenant_id},
        ["module_features"],
    )["module_features"].to_list()
    logging.debug("Raw user features fetched: %s", user_features_raw)

    # Parse the JSON string to a dictionary
    user_features = json.loads(
        user_features_raw[0]
    )  # Assuming it's a list with one JSON string

    # Initialize a list to hold features for the specified feature name
    features_list = []

    # Loop through all modules to find the specified feature name
    for module, features in user_features.items():
        if feature_name in features:
            features_list.extend(features[feature_name])
    logging.info("Retrieved features: %s", features_list)

    return features_list


def get_charge_history_data(data):
    """
    This function retrieves report data by executing a query based
    on the provided module name and parameters,
    converting the results into a dictionary format.
    It logs the audit and error details to the database
    and returns the report data along with a success flag.
    """
    logging.info("Request Recieved")
    # Record the start time for performance measurement
    start_time = time.time()
    module_name = data.get("module_name", "")
    request_received_at = data.get("request_received_at", "")
    session_id = data.get("sessionID", "")
    username = data.get("username", None)
    Partner = data.get("Partner", "")
    mod_pages = data.get("mod_pages", {})
    role_name = data.get("role_name", "")
    limit = 100
    offset = mod_pages.get("start", 0)
    end = mod_pages.get("end", 100)
    start = mod_pages.get("start", 0)
    table = data.get("table_name", "vw_optimization_smi_result_customer_charge_queue")
    # Database connection
    tenant_database = data.get("db_name", "")
    try:
        database = DB(tenant_database, **db_config)
        common_utils_database = DB(os.environ["COMMON_UTILS_DATABASE"], **db_config)
    except Exception as db_exception:
        logging.error("Failed to connect to the database: %s", str(db_exception))
        return {"error": "Database connection failed.", "details": str(db_exception)}
    return_json_data = {}
    # Initialize pagination dictionary
    pages = {"start": offset, "end": end}
    tenant_timezone = None
    try:
        count_params = [table]
        count_query = f"SELECT COUNT(*) FROM {table}"
        count_result = database.execute_query(count_query, count_params).iloc[0, 0]
        # Set total pages count
        pages["total"] = int(count_result)

        module_query_df = common_utils_database.get_data(
            "module_view_queries", {"module_name": module_name}
        )
        if module_query_df.empty:
            raise ValueError(f"No query found for module name: {module_name}")
        query = module_query_df.iloc[0]["module_query"]
        if not query:
            raise ValueError(f"Unknown module name: {module_name}")
        params = [offset, limit]
        # main_query_start_time = time.time()
        charge_history_df = database.execute_query(query, params=params)
        tenant_time_zone = fetch_tenant_timezone(data,common_utils_database)
        # conversion_start_time = time.time()
        # Convert timestamps
        charge_history_data = convert_timestamp_data(
            charge_history_df, tenant_time_zone
        )
        charge_history_data = charge_history_df.to_dict(orient="records")

        # Get headers mapping
        headers_map = get_headers_mapping(
            tenant_database,
            [module_name],
            role_name,
            "username",
            "main_tenant_id",
            "sub_parent_module",
            "parent_module",
            data,
            common_utils_database,
        )
        # Preparing the response data
        return_json_data.update(
            {
                "message": "Successfully Fetched the Charges history data",
                "flag": True,
                "headers_map": headers_map,
                "customer_charges_data": charge_history_data,
                "pages": pages,
            }
        )
        try:
            # End time and audit logging
            end_time = time.time()
            time_consumed = f"{end_time - start_time:.4f}"
            time_consumed = int(float(time_consumed))
            audit_data_user_actions = {
                "service_name": "Charges history",
                "created_date": request_received_at,
                "created_by": username,
                "status": str(return_json_data["flag"]),
                "time_consumed_secs": time_consumed,
                "session_id": session_id,
                "tenant_name": Partner,
                "comments": "Charges history data",
                "module_name": module_name,
                "request_received_at": request_received_at,
            }
            common_utils_database.update_audit(
                audit_data_user_actions, "audit_user_actions"
            )
        except Exception as e:
            logging.warning(f"Got Exception while updating the Audit")
        return return_json_data

    except Exception as e:
        logging.exception(f"An error occurred: {e}")
        # Get headers mapping
        headers_map = get_headers_mapping(
            tenant_database,
            [module_name],
            role_name,
            "username",
            "main_tenant_id",
            "sub_parent_module",
            "parent_module",
            data,
            common_utils_database,
        )
        message = f"Unable to fetch the Charges history data{e}"
        response = {"flag": True, "message": message, "headers_map": headers_map}
        error_type = str(type(e).__name__)
        error_data = {
            "service_name": "Charges history",
            "created_date": request_received_at,
            "error_message": message,
            "error_type": error_type,
            "users": username,
            "session_id": session_id,
            "tenant_name": Partner,
            "comments": "Charges history data",
            "module_name": module_name,
            "request_received_at": request_received_at,
        }
        common_utils_database.log_error_to_db(error_data, "error_log_table")
        return response


def fetch_tenant_timezone(data,common_utils_database):
    # Retrieve the tenant name from the provided data dictionary.
    tenant_name = data.get("tenant_name", "")
    # SQL query to fetch the time zone of the tenant from the 'tenant' table.
    tenant_timezone_query = (
        """SELECT time_zone FROM tenant WHERE tenant_name = %s"""
    )
    # Execute the query and retrieve the result from the database.
    tenant_timezone = common_utils_database.execute_query(
        tenant_timezone_query, params=[tenant_name]
    )
    if tenant_timezone.empty or tenant_timezone.iloc[0]["time_zone"] is None:
        raise ValueError("No valid timezone found for tenant.")
    tenant_time_zone = tenant_timezone.iloc[0]["time_zone"]
    match = re.search(
        r"\(\w+\s[+\-]?\d{2}:\d{2}:\d{2}\)\s*(Asia\s*/\s*Kolkata)",
        tenant_time_zone,
    )
    # If a match is found, format the time zone string by removing spaces.
    if match:
        tenant_time_zone = match.group(1).replace(
            " ", ""
        )  # Ensure it's formatted correctly
    return tenant_time_zone


def format_timestamp(ts):
    # Check if the timestamp is not None
    if ts is not None:
        # Convert a Timestamp or datetime object to the desired string format
        return ts.strftime("%b %d, %Y, %I:%M %p")
    else:
        # Return a placeholder or empty string if the timestamp is None
        return " "


def export_row_data_customer_charges(data):
    """
    Description:Exports data into an Excel file. It retrieves data based on the module
    name from the database,
    processes it, and returns a blob representation of the data if within the allowed row limit.
    """
    logging.info(f"Request Data Recieved")
    ### Extract parameters from the Request Data
    Partner = data.get("Partner", "")
    request_received_at = data.get("request_received_at", None)
    module_name = data.get("module_name", "Optimization row data")
    user_name = data.get("user_name", "")
    queue_id = data.get("queue_id", "")
    session_id = data.get("sessionID", "")
    ##database connection for common utilss
    try:
        ##databse connenction
        tenant_database = data.get("db_name", "")
        # database Connection
        database = DB(tenant_database, **db_config)
        db = DB(os.environ["COMMON_UTILS_DATABASE"], **db_config)
    except Exception as db_exception:
        logging.error("Failed to connect to the database: %s", str(db_exception))
        return {"error": "Database connection failed.", "details": str(db_exception)}
    # Start time  and date calculation
    start_time = time.time()
    try:

        # Fetch the query from the database based on the module name
        module_query_df = db.get_data(
            "export_queries", {"module_name": "Optimization row data"}
        )
        logging.info(module_query_df, "module_query_df")
        ##checking the dataframe is empty or not
        if module_query_df.empty:
            return {
                "flag": False,
                "message": f"No query found for module name: {module_name}",
            }
        # Extract the query string from the DataFrame
        query = module_query_df.iloc[0]["module_query"]
        if not query:
            return {"flag": False, "message": f"Unknown module name: {module_name}"}
        ##params for the specific module
        params = [queue_id]
        ##executing the query
        data_frame = database.execute_query(query, params=params)
        # Capitalize each word and add spaces
        data_frame.columns = [
            col.replace("_", " ").title() for col in data_frame.columns
        ]
        data_frame["S.NO"] = range(1, len(data_frame) + 1)
        # Reorder columns dynamically to put S.NO at the first position
        columns = ["S.NO"] + [col for col in data_frame.columns if col != "S.NO"]
        data_frame = data_frame[columns]
        # Proceed with the export if row count is within the allowed limit
        data_frame = data_frame.astype(str)
        data_frame.replace(to_replace="None", value="", inplace=True)

        blob_data = dataframe_to_blob(data_frame)
        # End time calculation
        end_time = time.time()
        time_consumed = f"{end_time - start_time:.4f}"
        time_consumed = int(float(time_consumed))
        # Create the filename with the date format YYYYMMDDHHMM
        current_time = datetime.now()
        formatted_date = current_time.strftime("%Y%m%d%H%M")
        filename = f"CustomerChargeDetailOutboardRecycle_{formatted_date}"
        # Return JSON response
        response = {
            "flag": True,
            "blob": blob_data.decode("utf-8"),
            "filename": filename,
        }
        audit_data_user_actions = {
            "service_name": "Module Management",
            "session_id": session_id,
            "created_date": request_received_at,
            "created_by": user_name,
            "status": str(response["flag"]),
            "time_consumed_secs": time_consumed,
            "tenant_name": Partner,
            "comments": "",
            "module_name": "export",
            "request_received_at": request_received_at,
        }
        db.update_audit(audit_data_user_actions, "audit_user_actions")
        return response
    except Exception as e:
        error_type = str(type(e).__name__)
        logging.exception(f"An error occurred: {e}")
        message = f"Error is {e}"
        response = {"flag": False, "message": message}
        try:
            # Log error to database
            error_data = {
                "service_name": "Module Management",
                "created_date": request_received_at,
                "error_message": message,
                "session_id": session_id,
                "error_type": error_type,
                "users": user_name,
                "tenant_name": Partner,
                "comments": message,
                "module_name": "export",
                "request_received_at": request_received_at,
            }
            db.log_error_to_db(error_data, "error_log_table")
        except Exception as e:
            logging.exception(f"Exception is {e}")
        return response


def dataframe_to_blob(data_frame):
    """
    Description:The Function is used to convert the dataframe to blob
    """
    # Create a BytesIO buffer
    bio = BytesIO()

    # Use ExcelWriter within a context manager to ensure proper saving
    with pd.ExcelWriter(bio, engine="openpyxl") as writer:
        data_frame.to_excel(writer, index=False)

    # Get the value from the buffer
    bio.seek(0)
    blob_data = base64.b64encode(bio.read())
    return blob_data


##charge history dropdown data
def customers_sessions_customer_charges_export_dropdown_data(data):
    logging.info("Request Data Received")

    # Extract parameters from the request data
    tenant_database = data.get("db_name", "")
    if not tenant_database:
        logging.error("Database name is missing from the request data.")
        return {
            "flag": False,
            "customer_sessions": {},
            "message": "Database name is required",
        }

    # Database connection setup
    try:
        database = DB(
            tenant_database, **db_config
        )  # Assuming DB is a predefined database connection class
    except Exception as e:
        logging.exception(f"Error establishing database connection: {e}")
        return {
            "flag": False,
            "customer_sessions": {},
            "message": "Error establishing database connection",
        }

    try:
        # Fetch customer names, session IDs, and billing period end date
        query = """
        SELECT DISTINCT customer_name, CAST(session_id AS TEXT) AS session_id,
        TO_CHAR(billing_period_end_date::date, 'YYYY-MM-DD') AS billing_period_end_date
        FROM vw_optimization_smi_result_customer_charge_queue;
        """

        # Get data from the database
        customer_session_df = database.execute_query(
            query, True
        )  # Assuming execute_query returns a DataFrame
        if customer_session_df.empty:
            logging.info("No data found for customer sessions.")
            return {
                "flag": True,
                "customer_sessions": {},
                "message": "No customer sessions found",
            }

        # Initialize a dictionary to store the result with differentiation of session details
        customer_sessions = {}

        # Iterate over the dataframe to build the dictionary
        for index, row in customer_session_df.iterrows():
            customer_name = row["customer_name"]
            session_id = row["session_id"]
            billing_period_end_date = row["billing_period_end_date"]

            # Check if the customer_name already exists in the dictionary
            if customer_name in customer_sessions:
                # Append a dictionary with session details for the existing customer
                customer_sessions[customer_name].append(
                    {
                        "session_id": session_id,
                        "billing_period_end_date": billing_period_end_date,
                    }
                )
            else:
                # Create a new list for the customer with session details
                customer_sessions[customer_name] = [
                    {
                        "session_id": session_id,
                        "billing_period_end_date": billing_period_end_date,
                    }
                ]

        response = {"flag": True, "customer_sessions": customer_sessions}
        return response

    except Exception as e:
        logging.exception(f"Error fetching data: {e}")
        return {"flag": False, "customer_sessions": {}, "message": f"Error: {str(e)}"}


def export_customer_charges(data):
    """
    Export customer charges to Excel files, zip them, and upload the zip file to an S3 bucket.

    Args:
        data (dict): Input data containing session_ids, db_name, and other details.

    Returns:
        dict: Response with a flag, message, and download_url if successful.
    """
    logging.info("Request Data Received")
    # Record the start time for performance measurement
    start_time = time.time()
    session_ids = data.get("session_ids", [])
    sessionID = data.get('sessionID', '')
    tenant_database = data.get("db_name", "")
    request_received_at = data.get("request_received_at", "")
    username = data.get("username", "")
    module_name = data.get("module_name", "")
    partner = data.get("Partner", "")
    database = DB(
        tenant_database, **db_config
    )  # Assuming DB class is defined elsewhere
    zip_filename = "CustomerCharges.zip"
    s3_key = f"exports/customer_charges/{zip_filename}"
    db = DB(os.environ["COMMON_UTILS_DATABASE"], **db_config)

    # Update export status to 'Waiting'
    db.update_dict(
        "export_status",
        {"status_flag": "Waiting", "url": ""},
        {"module_name": "Charge History"},
    )
    try:
        # Create an in-memory buffer to hold the zip file
        zip_buffer = io.BytesIO()

        with zipfile.ZipFile(zip_buffer, "w", zipfile.ZIP_DEFLATED) as zipf:
                try:
                    logging.info(f"Processing all customer charges")
                    query = """
                        SELECT rev_account_number, customer_name,
                            billing_period_end_date - billing_period_start_date AS billing_period_duration,
                            billing_period_start_date, billing_period_end_date, usage_mb,
                            base_charge_amount, rate_charge_amt, overage_charge_amount,
                            is_processed, error_message,
                            base_charge_amount + rate_charge_amt + overage_charge_amount AS total_data_charge_amount,
                            iccid, msisdn, sms_usage, sms_charge_amount, total_charge_amount
                        FROM vw_optimization_smi_result_customer_charge_queue
                    """
                    dataframe = database.execute_query(query,True)

                    # Validate DataFrame content
                    if dataframe.empty:
                        logging.warning(f"No data found ")
                        return {"flag": False, "message": "No data found"}

                    logging.debug(
                        f"DataFrame shape: {dataframe.shape}"
                    )

                    # Rename columns for better readability
                    dataframe.columns = [
                        col.replace("_", " ").title() for col in dataframe.columns
                    ]
                    dataframe["S.NO"] = range(1, len(dataframe) + 1)
                    columns = ["S.NO"] + [
                        col for col in dataframe.columns if col != "S.NO"
                    ]
                    dataframe = dataframe[columns]

                    # Extract billing period dates
                    billing_period_start_date = dataframe.iloc[0][
                        "Billing Period Start Date"
                    ].strftime("%Y%m%d")
                    billing_period_end_date = dataframe.iloc[0][
                        "Billing Period End Date"
                    ].strftime("%Y%m%d")

                    logging.debug(
                        f"Billing Period: {billing_period_start_date} to {billing_period_end_date}"
                    )

                    # Create filename for Excel
                    excel_filename = f"CustomerChargeDetail_{billing_period_start_date}_{billing_period_end_date}.xlsx"

                    # Write DataFrame to an in-memory Excel file
                    customer_excel_buffer = io.BytesIO()
                    with pd.ExcelWriter(
                        customer_excel_buffer, engine="openpyxl"
                    ) as writer:
                        dataframe.to_excel(writer, index=False, sheet_name="Charges")

                        # Apply formatting
                        workbook = writer.book
                        sheet = writer.sheets["Charges"]

                        # Center align header cells and add background color
                        for cell in sheet[1]:
                            cell.alignment = Alignment(horizontal="center", vertical="center")
                            cell.fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")

                        # Adjust column widths based on content
                        for col_idx, col_cells in enumerate(sheet.columns, 1):
                            max_length = max(len(str(cell.value or "")) for cell in col_cells)
                            sheet.column_dimensions[get_column_letter(col_idx)].width = max_length + 2

                    # Write Excel file to the zip archive
                    customer_excel_buffer.seek(0)
                    zipf.writestr(excel_filename, customer_excel_buffer.getvalue())
                    customer_excel_buffer.close()

                except Exception as e:
                    logging.exception(f"Error processing charges: {e}")

        # Upload the zip file to S3
        zip_buffer.seek(0)
        try:
            s3_client = boto3.client("s3")
            s3_client.put_object(
                Bucket=S3_BUCKET_NAME,
                Key=s3_key,
                Body=zip_buffer.getvalue(),
                ContentType="application/zip",
            )
            download_url = f"https://{S3_BUCKET_NAME}.s3.amazonaws.com/{s3_key}"
            logging.info("File uploaded to S3 successfully.")
            response = {
                "flag": True,
                "message": "Export successful",
                "download_url": download_url,
            }
            db.update_dict(
                "export_status",
                {"status_flag": "Success", "url": download_url},
                {"module_name": "Charge History"},
            )
            # End time and audit logging
            end_time = time.time()
            time_consumed = f"{end_time - start_time:.4f}"
            time_consumed = int(float(time_consumed))
            audit_data_user_actions = {
                "service_name": "charges_history",
                "created_date": request_received_at,
                "created_by": username,
                "status": str(response["flag"]),
                "time_consumed_secs": time_consumed,
                "session_id": sessionID,
                "tenant_name": partner,
                "comments": "Uploading to S3 bucket",
                "module_name": "",
                "request_received_at": request_received_at,
            }
            db.update_audit(audit_data_user_actions, "audit_user_actions")
            return response
        except Exception as e:
            logging.exception(f"Error uploading to S3: {e}")
            return {"flag": False, "message": "Failed to upload zip file to S3"}
    except Exception as e:
        logging.exception(f"Exceptions is {e}")
        message = f"Failed to upload zip file to S3"
        return_json_data = {"flag": True, "message": message}
        error_type = str(type(e).__name__)
        # Error logging
        error_data = {
            "service_name": "charges_history",
            "created_date": request_received_at,
            "error_message": message,
            "error_type": error_type,
            "users": username,
            "session_id": sessionID,
            "tenant_name": partner,
            "comments": message,
            "module_name": module_name,
            "request_received_at": request_received_at,
        }
        db.log_error_to_db(error_data, "error_log_table")

        return return_json_data


def convert_timestamp_data(df, tenant_time_zone):
    """Convert timestamp columns in the DataFrame to the tenant's timezone."""
    # List of timestamp columns to convert
    timestamp_columns = [
        "created_date",
        "modified_date",
        "deleted_date",
    ]  # Adjust as needed based on your data

    # Convert specified timestamp columns to the tenant's timezone
    for col in timestamp_columns:
        if col in df.columns:
            # Convert the column to datetime format if it isn't already
            df[col] = pd.to_datetime(df[col], errors="coerce")
            # Localize the datetime column to UTC if it's naive
            df[col] = df[col].dt.tz_localize("UTC", ambiguous="NaT", nonexistent="NaT")
            # Convert to the target timezone
            df[col] = df[col].dt.tz_convert(tenant_time_zone)
            # Format as string if needed (optional)
            df[col] = df[col].dt.strftime("%m-%d-%Y %H:%M:%S")

    return df


def serialize_data(data):
    """Recursively convert pandas objects in the data structure to serializable types."""
    if isinstance(data, list):
        return [serialize_data(item) for item in data]
    elif isinstance(data, dict):
        return {key: serialize_data(value) for key, value in data.items()}
    elif isinstance(data, pd.Timestamp):
        return data.strftime("%m-%d-%Y %H:%M:%S")  # Convert to string
    else:
        return data  # Return as is if not a pandas object


def get_export_status(data):
    try:
        common_utils_database = DB("common_utils", **db_config)
        module_name = data.get("module_name")
        export_status_query = (
            f"select * from export_status where module_name='{module_name}'"
        )
        export_status_data = common_utils_database.execute_query(
            export_status_query, True
        ).to_dict(orient="records")
        return {"flag": True, "export_status_data": export_status_data[0]}
    except Exception as e:
        logging.exception(f"Exception is {e}")
        return {"flag": False, "export_status_data": []}
